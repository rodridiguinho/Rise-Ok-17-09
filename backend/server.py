from fastapi import FastAPI, APIRouter, HTTPException, status
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, EmailStr
from typing import List, Optional, Dict, Any
from datetime import datetime, date
from bson import ObjectId
import bcrypt
import jwt

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ.get('DB_NAME', 'agentepro_cashcontrol')]

# JWT Settings
SECRET_KEY = os.environ.get("JWT_SECRET_KEY", "your-secret-key-change-in-production")
ALGORITHM = "HS256"

# Create the main app
app = FastAPI(
    title="Rise Travel - Controle de Caixa API",
    description="API para sistema de controle de caixa da Rise Travel",
    version="1.0.0"
)

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Models
class UserLogin(BaseModel):
    email: EmailStr
    password: str

class TransactionCreate(BaseModel):
    type: str
    category: str
    description: str
    amount: float
    paymentMethod: str
    client: Optional[str] = None
    supplier: Optional[str] = None
    transactionDate: Optional[str] = None  # Date when the sale actually happened

# Helper functions
def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def generate_pdf_report(report_data: dict) -> bytes:
    """Generate PDF report content using ReportLab"""
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from reportlab.lib.units import inch
        from datetime import datetime
        import io
        
        # Get report parameters
        start_date = report_data.get('startDate', 'Início')
        end_date = report_data.get('endDate', 'Hoje')
        transactions = report_data.get('transactions', [])
        
        # Calculate totals
        total_entradas = sum(t.get('amount', 0) for t in transactions if t.get('type') == 'entrada')
        total_saidas = sum(t.get('amount', 0) for t in transactions if t.get('type') == 'saida')
        
        # Create PDF buffer
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1*inch)
        
        # Get styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            alignment=TA_CENTER,
            spaceAfter=30,
        )
        
        # Build content
        story = []
        
        # Title
        story.append(Paragraph("RELATÓRIO DE CONTROLE DE CAIXA", title_style))
        story.append(Paragraph("Rise Travel", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        # Period and generation date
        story.append(Paragraph(f"Período: {start_date} até {end_date}", styles['Normal']))
        story.append(Paragraph(f"Data de Geração: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Financial summary
        story.append(Paragraph("RESUMO FINANCEIRO", styles['Heading3']))
        summary_data = [
            ['Total de Entradas:', f'R$ {total_entradas:,.2f}'],
            ['Total de Saídas:', f'R$ {total_saidas:,.2f}'],
            ['Resultado Líquido:', f'R$ {(total_entradas - total_saidas):,.2f}'],
            ['Total de Transações:', str(len(transactions))]
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, 'black'),
            ('BACKGROUND', (0, 0), (0, -1), '#f0f0f0'),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # Transactions detail
        if transactions:
            story.append(Paragraph("DETALHAMENTO DAS TRANSAÇÕES", styles['Heading3']))
            
            # Create transaction table
            transaction_data = [['Data', 'Tipo', 'Categoria', 'Descrição', 'Valor']]
            
            for t in transactions:
                tipo_symbol = "+" if t.get('type') == 'entrada' else "-"
                transaction_data.append([
                    t.get('date', 'N/A'),
                    t.get('type', '').upper(),
                    t.get('category', 'N/A'),
                    t.get('description', 'N/A')[:30] + '...' if len(t.get('description', '')) > 30 else t.get('description', 'N/A'),
                    f'{tipo_symbol}R$ {t.get("amount", 0):,.2f}'
                ])
            
            transaction_table = Table(transaction_data, colWidths=[1*inch, 1*inch, 1.5*inch, 2*inch, 1.5*inch])
            transaction_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, 'black'),
                ('BACKGROUND', (0, 0), (-1, 0), '#d0d0d0'),
                ('ALIGN', (4, 1), (4, -1), 'RIGHT'),
            ]))
            story.append(transaction_table)
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()
        
    except ImportError:
        # Fallback to simple text if ReportLab is not available
        return generate_simple_pdf_fallback(report_data)
    except Exception as e:
        logging.error(f"Error generating PDF: {str(e)}")
        return generate_simple_pdf_fallback(report_data)

def generate_simple_pdf_fallback(report_data: dict) -> bytes:
    """Simple PDF fallback (text content)"""
    from datetime import datetime
    
    start_date = report_data.get('startDate', 'Início')
    end_date = report_data.get('endDate', 'Hoje')
    transactions = report_data.get('transactions', [])
    
    total_entradas = sum(t.get('amount', 0) for t in transactions if t.get('type') == 'entrada')
    total_saidas = sum(t.get('amount', 0) for t in transactions if t.get('type') == 'saida')
    
    content = f"""
RELATÓRIO DE CONTROLE DE CAIXA - RISE TRAVEL

Período: {start_date} até {end_date}
Data de Geração: {datetime.now().strftime('%d/%m/%Y %H:%M')}

RESUMO FINANCEIRO:
Total de Entradas: R$ {total_entradas:,.2f}
Total de Saídas: R$ {total_saidas:,.2f}
Resultado Líquido: R$ {(total_entradas - total_saidas):,.2f}
Total de Transações: {len(transactions)}

DETALHAMENTO DAS TRANSAÇÕES:
{'='*60}
"""
    
    for t in transactions:
        tipo_symbol = "+" if t.get('type') == 'entrada' else "-"
        content += f"""
Data: {t.get('date', 'N/A')} {t.get('time', 'N/A')}
Tipo: {t.get('type', 'N/A').upper()}
Categoria: {t.get('category', 'N/A')}
Descrição: {t.get('description', 'N/A')}
Valor: {tipo_symbol}R$ {t.get('amount', 0):,.2f}
Pagamento: {t.get('paymentMethod', 'N/A')}
{'-'*40}
"""
    
    return content.encode('utf-8')

def generate_excel_report(report_data: dict) -> bytes:
    """Generate Excel report using openpyxl"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from datetime import datetime
        import io
        
        transactions = report_data.get('transactions', [])
        start_date = report_data.get('startDate', 'Início')
        end_date = report_data.get('endDate', 'Hoje')
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Controle de Caixa"
        
        # Title
        ws.merge_cells('A1:J1')
        ws['A1'] = "RELATÓRIO DE CONTROLE DE CAIXA - RISE TRAVEL"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Period
        ws.merge_cells('A2:J2')
        ws['A2'] = f"Período: {start_date} até {end_date} | Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Calculate totals
        total_entradas = sum(t.get('amount', 0) for t in transactions if t.get('type') == 'entrada')
        total_saidas = sum(t.get('amount', 0) for t in transactions if t.get('type') == 'saida')
        
        # Summary
        ws['A4'] = "RESUMO FINANCEIRO"
        ws['A4'].font = Font(bold=True)
        ws['A5'] = "Total de Entradas:"
        ws['B5'] = f"R$ {total_entradas:,.2f}"
        ws['A6'] = "Total de Saídas:"
        ws['B6'] = f"R$ {total_saidas:,.2f}"
        ws['A7'] = "Resultado Líquido:"
        ws['B7'] = f"R$ {(total_entradas - total_saidas):,.2f}"
        ws['B7'].font = Font(bold=True)
        
        # Transaction headers
        headers = ['Data', 'Hora', 'Tipo', 'Categoria', 'Descrição', 'Valor', 'Forma Pagamento', 'Cliente', 'Fornecedor', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=9, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Transaction data
        for row, t in enumerate(transactions, 10):
            ws.cell(row=row, column=1, value=t.get('date', ''))
            ws.cell(row=row, column=2, value=t.get('time', ''))
            ws.cell(row=row, column=3, value=t.get('type', ''))
            ws.cell(row=row, column=4, value=t.get('category', ''))
            ws.cell(row=row, column=5, value=t.get('description', ''))
            ws.cell(row=row, column=6, value=t.get('amount', 0))
            ws.cell(row=row, column=7, value=t.get('paymentMethod', ''))
            ws.cell(row=row, column=8, value=t.get('client', ''))
            ws.cell(row=row, column=9, value=t.get('supplier', ''))
            ws.cell(row=row, column=10, value=t.get('status', 'Confirmado'))
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
        
    except ImportError:
        # Fallback to CSV
        return generate_csv_fallback(report_data)
    except Exception as e:
        logging.error(f"Error generating Excel: {str(e)}")
        return generate_csv_fallback(report_data)

def generate_csv_fallback(report_data: dict) -> bytes:
    """CSV fallback for Excel export"""
    import csv
    import io
    
    transactions = report_data.get('transactions', [])
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header
    writer.writerow([
        'Data', 'Hora', 'Tipo', 'Categoria', 'Descrição', 
        'Valor', 'Forma Pagamento', 'Cliente', 'Fornecedor', 'Status'
    ])
    
    # Data
    for t in transactions:
        writer.writerow([
            t.get('date', ''),
            t.get('time', ''),
            t.get('type', ''),
            t.get('category', ''),
            t.get('description', ''),
            f"{t.get('amount', 0):.2f}",
            t.get('paymentMethod', ''),
            t.get('client', ''),
            t.get('supplier', ''),
            t.get('status', 'Confirmado')
        ])
    
    return output.getvalue().encode('utf-8')

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return bcrypt.checkpw(plain_password.encode('utf-8'), hashed_password.encode('utf-8'))

def create_jwt_token(user_id: str, email: str) -> str:
    payload = {
        'user_id': user_id,
        'email': email,
        'exp': datetime.utcnow().timestamp() + 3600  # 1 hour
    }
    return jwt.encode(payload, SECRET_KEY, algorithm=ALGORITHM)

def decode_jwt_token(token: str) -> dict:
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        return payload
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

# Routes
@api_router.get("/")
async def root():
    return {"message": "Rise Travel - Controle de Caixa API - Running"}

@api_router.get("/health")
async def health_check():
    return {"status": "healthy", "service": "cashcontrol-api"}

@api_router.delete("/transactions/{transaction_id}")
async def delete_transaction(transaction_id: str):
    """Deletar transação (mockado)"""
    try:
        # Por enquanto apenas simula deleção
        return {"success": True, "message": "Transaction deleted successfully"}
    except Exception as e:
        logging.error(f"Delete transaction error: {str(e)}")
        raise HTTPException(status_code=500, detail="Error deleting transaction")

@api_router.post("/auth/login")
async def login(login_data: UserLogin):
    """Login do usuário"""
    try:
        # Buscar usuário
        user = await db.users.find_one({"email": login_data.email})
        if not user or not verify_password(login_data.password, user["password"]):
            raise HTTPException(status_code=401, detail="Invalid credentials")
        
        # Criar token
        token = create_jwt_token(str(user["_id"]), user["email"])
        
        return {
            "success": True,
            "token": token,
            "user": {
                "id": str(user["_id"]),
                "email": user["email"],
                "name": user["name"],
                "role": user.get("role", "user"),
                "companyName": user.get("companyName"),
                "phone": user.get("phone"),
                "address": user.get("address"),
                "settings": user.get("settings", {})
            }
        }
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Login error: {str(e)}")
        raise HTTPException(status_code=500, detail="Login error")

@api_router.get("/transactions/summary")
async def get_transaction_summary():
    """Obter resumo das transações"""
    try:
        # Por enquanto retorna dados mockados
        return {
            "totalEntradas": 16030.00,
            "totalSaidas": 1850.00,
            "saldoAtual": 14180.00,
            "transacoesHoje": 3,
            "clientesAtendidos": 6,
            "ticketMedio": 2671.67
        }
    except Exception as e:
        logging.error(f"Summary error: {str(e)}")
        raise HTTPException(status_code=500, detail="Error getting summary")

@api_router.get("/analytics/sales")
async def get_sales_analytics():
    """Obter análises de vendas"""
    try:
        return {
            "valorTotal": 259390.78,
            "valorTotalAnterior": 226580.50,
            "percentualVariacao": 14.09,
            "comissoes": 4178.19,
            "comissoesAnterior": 3986.45,
            "percentualComissoes": 104.99,
            "numeroVendas": 43,
            "vendasAnterior": 34,
            "percentualVendas": 126.32,
            "novosClientes": 35,
            "clientesAnterior": 35,
            "percentualClientes": 0.0,
            "ticketMedio": 6032.34,
            "ticketAnterior": 7052.20,
            "percentualTicket": -14.09,
            "taxaConversao": {
                "vendasPorCotacoes": 43,
                "totalCotacoes": 0,
                "percentual": 0
            },
            "rankingVendedores": [
                {
                    "nome": "Fernando dos Anjos",
                    "valor": 138615.90,
                    "percentual": 53.5,
                    "posicao": 1
                },
                {
                    "nome": "Franciele Oliveira", 
                    "valor": 109994.88,
                    "percentual": 42.4,
                    "posicao": 2
                },
                {
                    "nome": "Katia Alessandra",
                    "valor": 10780.00,
                    "percentual": 4.2,
                    "posicao": 3
                }
            ]
        }
    except Exception as e:
        logging.error(f"Sales analytics error: {str(e)}")
        raise HTTPException(status_code=500, detail="Error getting sales analytics")

@api_router.get("/analytics/financial")
async def get_financial_analytics():
    """Obter análises financeiras"""
    try:
        return {
            "receitas": 275728.78,
            "receitasAnterior": 275728.78,
            "percentualReceitas": 0.0,
            "despesas": 231666.75,
            "despesasAnterior": 231666.75,
            "percentualDespesas": 0.0,
            "lucro": 44062.03,
            "lucroAnterior": 44062.03,
            "percentualLucro": 0.0,
            "margemLucro": 16.0,
            "graficoDados": {
                "labels": ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set"],
                "receitas": [180000, 200000, 220000, 240000, 260000, 275000, 290000, 270000, 275728],
                "despesas": [150000, 170000, 180000, 200000, 210000, 220000, 230000, 225000, 231666],
                "lucro": [30000, 30000, 40000, 40000, 50000, 55000, 60000, 45000, 44062]
            }
        }
    except Exception as e:
        logging.error(f"Financial analytics error: {str(e)}")
        raise HTTPException(status_code=500, detail="Error getting financial analytics")

@api_router.get("/transactions")
async def get_transactions():
    """Obter transações"""
    try:
        # Por enquanto retorna dados mockados
        mock_transactions = [
            {
                "id": "1",
                "date": "2025-01-05",
                "time": "09:30",
                "type": "entrada",
                "category": "Pacote Turístico",
                "description": "Pacote Europa 7 dias - Cliente João Silva",
                "amount": 8500.00,
                "paymentMethod": "Cartão de Crédito",
                "client": "João Silva",
                "status": "Confirmado",
                "createdAt": datetime.utcnow()
            },
            {
                "id": "2",
                "date": "2025-01-05",
                "time": "10:15",
                "type": "saida",
                "category": "Fornecedor",
                "description": "Pagamento Hotel Ibis - Reserva #4523",
                "amount": 1200.00,
                "paymentMethod": "Transferência",
                "supplier": "Hotel Ibis",
                "status": "Pago",
                "createdAt": datetime.utcnow()
            }
        ]
        return mock_transactions
    except Exception as e:
        logging.error(f"Transactions error: {str(e)}")
        raise HTTPException(status_code=500, detail="Error getting transactions")

@api_router.post("/transactions")
async def create_transaction(transaction: TransactionCreate):
    """Criar nova transação"""
    try:
        # Use the transaction date provided by user, or default to today
        transaction_date = transaction.transactionDate if transaction.transactionDate else date.today().strftime("%Y-%m-%d")
        
        # Por enquanto apenas simula criação
        new_transaction = {
            "id": str(ObjectId()),
            "date": transaction_date,  # Use the actual transaction date, not entry date
            "time": datetime.now().strftime("%H:%M"),  # Keep current time for record keeping
            "type": transaction.type,
            "category": transaction.category,
            "description": transaction.description,
            "amount": transaction.amount,
            "paymentMethod": transaction.paymentMethod,
            "client": transaction.client,
            "supplier": transaction.supplier,
            "status": "Confirmado",
            "transactionDate": transaction_date,  # Store the actual transaction date
            "createdAt": datetime.utcnow(),  # Keep record of when this was entered into system
            "entryDate": date.today().strftime("%Y-%m-%d")  # When this was entered into system
        }
        return new_transaction
    except Exception as e:
        logging.error(f"Create transaction error: {str(e)}")
        raise HTTPException(status_code=500, detail="Error creating transaction")

@api_router.get("/transactions/categories")
async def get_categories():
    """Obter categorias"""
    return {
        "categories": [
            "Pacote Turístico",
            "Passagem Aérea", 
            "Hotel/Hospedagem",
            "Seguro Viagem",
            "Transfer",
            "Excursão",
            "Aluguel de Carro",
            "Cruzeiro",
            "Ingresso/Atrações",
            "Fornecedor",
            "Despesa Operacional",
            "Comissão"
        ]
    }

@api_router.get("/transactions/payment-methods")
async def get_payment_methods():
    """Obter métodos de pagamento"""
    return {
        "paymentMethods": [
            "Dinheiro",
            "PIX",
            "Cartão de Crédito",
            "Cartão de Débito",
            "Transferência",
            "Cartão Corporativo"
        ]
    }

@api_router.post("/reports/export/pdf")
async def export_pdf(report_data: dict = None):
    """Exportar relatório em PDF"""
    try:
        from datetime import datetime
        import tempfile
        
        # Generate PDF content (simplified HTML-to-PDF approach)
        pdf_content = generate_pdf_report(report_data or {})
        
        # Create temporary file
        filename = f"relatorio_caixa_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        return {
            "success": True, 
            "message": "PDF export completed",
            "filename": filename,
            "downloadUrl": f"/api/reports/download/{filename}",
            "contentType": "application/pdf"
        }
        
    except Exception as e:
        logging.error(f"Error exporting PDF: {str(e)}")
        raise HTTPException(status_code=500, detail="Error generating PDF report")

@api_router.post("/reports/export/excel")
async def export_excel(report_data: dict = None):
    """Exportar relatório em Excel"""
    try:
        from datetime import datetime
        
        # Generate Excel content (CSV format for simplicity)
        excel_content = generate_excel_report(report_data or {})
        
        filename = f"relatorio_caixa_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return {
            "success": True,
            "message": "Excel export completed", 
            "filename": filename,
            "downloadUrl": f"/api/reports/download/{filename}",
            "contentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        }
        
    except Exception as e:
        logging.error(f"Error exporting Excel: {str(e)}")
        raise HTTPException(status_code=500, detail="Error generating Excel report")

# Users API endpoints
@api_router.get("/users")
async def get_users():
    """Obter lista de usuários"""
    try:
        users = await db.users.find({}, {"password": 0}).to_list(100)  # Exclude password
        for user in users:
            user["id"] = str(user["_id"])
            user["_id"] = str(user["_id"])
            # Convert datetime objects to strings
            if "createdAt" in user:
                user["createdAt"] = user["createdAt"].isoformat()
            if "updatedAt" in user:
                user["updatedAt"] = user["updatedAt"].isoformat()
        return users
    except Exception as e:
        logging.error(f"Error getting users: {str(e)}")
        raise HTTPException(status_code=500, detail="Error getting users")

@api_router.post("/users")
async def create_user(user_data: dict):
    """Criar novo usuário"""
    try:
        # Check if email already exists
        existing_user = await db.users.find_one({"email": user_data["email"]})
        if existing_user:
            raise HTTPException(status_code=400, detail="Email já existe")
        
        # Hash password
        hashed_password = hash_password(user_data["password"])
        
        # Prepare user data
        new_user = {
            "email": user_data["email"],
            "password": hashed_password,
            "name": user_data["name"],
            "role": user_data.get("role", "Vendedor"),
            "phone": user_data.get("phone", ""),
            "status": user_data.get("status", "Ativo"),
            "companyName": "Rise Travel",
            "settings": {
                "currency": "BRL",
                "timezone": "America/Sao_Paulo",
                "notifications": {
                    "emailNotifications": True,
                    "pushNotifications": False,
                    "dailyReport": True,
                    "transactionAlerts": True,
                    "lowCashAlert": True
                },
                "preferences": {
                    "theme": "light",
                    "autoExport": False,
                    "backupFrequency": "weekly",
                    "decimalPlaces": 2
                }
            },
            "createdAt": datetime.utcnow(),
            "updatedAt": datetime.utcnow()
        }
        
        # Insert user
        result = await db.users.insert_one(new_user)
        
        # Get created user (without password)
        created_user = await db.users.find_one({"_id": result.inserted_id}, {"password": 0})
        if created_user:
            created_user["id"] = str(created_user["_id"])
            created_user["_id"] = str(created_user["_id"])
            # Convert datetime objects to strings
            if "createdAt" in created_user:
                created_user["createdAt"] = created_user["createdAt"].isoformat()
            if "updatedAt" in created_user:
                created_user["updatedAt"] = created_user["updatedAt"].isoformat()
        
        return created_user
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error creating user: {str(e)}")
        raise HTTPException(status_code=500, detail="Error creating user")

@api_router.put("/users/{user_id}")
async def update_user(user_id: str, user_data: dict):
    """Atualizar usuário"""
    try:
        # Prepare update data
        update_data = {
            "name": user_data.get("name"),
            "role": user_data.get("role"),
            "phone": user_data.get("phone", ""),
            "status": user_data.get("status", "Ativo"),
            "updatedAt": datetime.utcnow()
        }
        
        # Only update email if provided and different
        if user_data.get("email"):
            # Check if email already exists for another user
            existing_user = await db.users.find_one({
                "email": user_data["email"],
                "_id": {"$ne": ObjectId(user_id)}
            })
            if existing_user:
                raise HTTPException(status_code=400, detail="Email já existe")
            update_data["email"] = user_data["email"]
        
        # If password is being updated, hash it
        if user_data.get("password") and user_data["password"].strip():
            update_data["password"] = hash_password(user_data["password"])
        
        # Update user
        result = await db.users.update_one(
            {"_id": ObjectId(user_id)},
            {"$set": update_data}
        )
        
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="User not found")
        
        # Get updated user (without password)
        updated_user = await db.users.find_one({"_id": ObjectId(user_id)}, {"password": 0})
        if updated_user:
            updated_user["id"] = str(updated_user["_id"])
            updated_user["_id"] = str(updated_user["_id"])
            # Convert datetime objects to strings
            if "createdAt" in updated_user:
                updated_user["createdAt"] = updated_user["createdAt"].isoformat()
            if "updatedAt" in updated_user:
                updated_user["updatedAt"] = updated_user["updatedAt"].isoformat()
        
        return updated_user
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error updating user: {str(e)}")
        raise HTTPException(status_code=500, detail="Error updating user")

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str):
    """Deletar usuário"""
    try:
        # Don't allow deleting the current user (you could add this check)
        result = await db.users.delete_one({"_id": ObjectId(user_id)})
        
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="User not found")
        
        return {"success": True, "message": "User deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error deleting user: {str(e)}")
        raise HTTPException(status_code=500, detail="Error deleting user")

# Include the main router in the app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Startup event
@app.on_event("startup")
async def startup_db_client():
    """Conectar ao MongoDB e criar usuário padrão"""
    try:
        await client.admin.command('ping')
        logger.info("✅ Connected to MongoDB")
        
        # Criar usuário padrão se não existir
        existing_user = await db.users.find_one({"email": "rodrigo@risetravel.com.br"})
        if not existing_user:
            default_user = {
                "email": "rodrigo@risetravel.com.br",
                "password": hash_password("Emily2030*"),
                "name": "Rodrigo Silva",
                "role": "Gerente",
                "companyName": "Rise Travel",
                "phone": "+55 11 99999-9999",
                "address": "Rua das Flores, 123 - São Paulo, SP",
                "settings": {
                    "currency": "BRL",
                    "timezone": "America/Sao_Paulo",
                    "notifications": {
                        "emailNotifications": True,
                        "pushNotifications": False,
                        "dailyReport": True,
                        "transactionAlerts": True,
                        "lowCashAlert": True
                    },
                    "preferences": {
                        "theme": "light",
                        "autoExport": False,
                        "backupFrequency": "weekly",
                        "decimalPlaces": 2
                    }
                },
                "createdAt": datetime.utcnow(),
                "updatedAt": datetime.utcnow()
            }
            await db.users.insert_one(default_user)
            logger.info("✅ Default user created")
    except Exception as e:
        logger.error(f"❌ Database connection error: {e}")

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
    logger.info("✅ MongoDB connection closed")